import sys
import sqlite3
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QVBoxLayout, QLineEdit, QMessageBox
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt

class CelengankuGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.db_path = "finance.db"
        self.excel_path = "riwayat_transaksi.xlsx"
        self.log_path = "logs.txt"
        self.setup_database()
        self.setup_excel()
        self.setup_logger()
        self.balance = self.get_balance()
        self.initUI()

    def setup_database(self):
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS keuangan (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        tipe TEXT,
                        jumlah INTEGER,
                        waktu TEXT
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS saldo (
                        id INTEGER PRIMARY KEY CHECK (id = 1),
                        jumlah INTEGER
                    )''')
        c.execute("INSERT OR IGNORE INTO saldo (id, jumlah) VALUES (1, 0)")
        conn.commit()
        conn.close()

    def setup_excel(self):
        try:
            load_workbook(self.excel_path)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Waktu", "Tipe", "Jumlah", "Saldo"])
            wb.save(self.excel_path)

    def setup_logger(self):
        logging.basicConfig(filename=self.log_path, level=logging.INFO, format='%(asctime)s | %(message)s')

    def get_balance(self):
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute("SELECT jumlah FROM saldo WHERE id = 1")
        result = c.fetchone()
        conn.close()
        return result[0] if result else 0

    def update_balance(self):
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute("UPDATE saldo SET jumlah = ? WHERE id = 1", (self.balance,))
        conn.commit()
        conn.close()

    def save_transaction(self, tipe, jumlah):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute("INSERT INTO keuangan (tipe, jumlah, waktu) VALUES (?, ?, ?)", (tipe, jumlah, now))
        conn.commit()
        conn.close()

        wb = load_workbook(self.excel_path)
        ws = wb.active
        ws.append([now, tipe, jumlah, self.balance])
        wb.save(self.excel_path)

        logging.info(f"{tipe} | Rp{jumlah} | Saldo: Rp{self.balance}")

    def initUI(self):
        self.setWindowTitle("Celenganku - Galaxy Edition 🚀")
        self.setGeometry(100, 100, 500, 400)
        self.setStyleSheet("background-color: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #0d0d2b, stop:1 #1a1a40); color: white;")

        font = QFont("Arial", 12)
        title_font = QFont("Orbitron", 16, QFont.Bold)

        self.label = QLabel("💫 CELEGANKU - GALAXY EDITION 💫")
        self.label.setFont(title_font)
        self.label.setAlignment(Qt.AlignCenter)

        self.balance_label = QLabel(f"Saldo Galaksi: Rp{self.balance}")
        self.balance_label.setFont(font)
        self.balance_label.setAlignment(Qt.AlignCenter)

        self.amount_input = QLineEdit()
        self.amount_input.setPlaceholderText("Masukkan jumlah...")
        self.amount_input.setFont(font)
        self.amount_input.setStyleSheet("background-color: #33334d; color: white; border-radius: 10px; padding: 5px;")

        deposit_btn = QPushButton("🌟 Tambah Tabungan")
        deposit_btn.clicked.connect(self.deposit)

        withdraw_btn = QPushButton("🌌 Tarik Uang")
        withdraw_btn.clicked.connect(self.withdraw)

        exit_btn = QPushButton("🚀 Keluar")
        exit_btn.clicked.connect(self.close)

        for btn in [deposit_btn, withdraw_btn, exit_btn]:
            btn.setFont(font)
            btn.setStyleSheet("background-color: #29294d; color: white; border-radius: 10px; padding: 10px;")

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.balance_label)
        layout.addWidget(self.amount_input)
        layout.addWidget(deposit_btn)
        layout.addWidget(withdraw_btn)
        layout.addWidget(exit_btn)
        self.setLayout(layout)

    def deposit(self):
        try:
            amount = int(self.amount_input.text())
            if amount > 0:
                self.balance += amount
                self.update_balance()
                self.save_transaction("Tabung", amount)
                self.balance_label.setText(f"Saldo Galaksi: Rp{self.balance}")
                QMessageBox.information(self, "Berhasil", f"💰 Berhasil menyimpan Rp{amount} ke tabungan galaksi!")
            else:
                QMessageBox.warning(self, "Error", "Jumlah harus lebih besar dari 0.")
        except ValueError:
            QMessageBox.warning(self, "Error", "Masukkan angka yang valid.")

    def withdraw(self):
        try:
            amount = int(self.amount_input.text())
            if 0 < amount <= self.balance:
                self.balance -= amount
                self.update_balance()
                self.save_transaction("Tarik", amount)
                self.balance_label.setText(f"Saldo Galaksi: Rp{self.balance}")
                QMessageBox.information(self, "Berhasil", f"🌌 Berhasil menarik Rp{amount} dari tabungan galaksi!")
            else:
                QMessageBox.warning(self, "Error", "Saldo tidak mencukupi atau jumlah tidak valid.")
        except ValueError:
            QMessageBox.warning(self, "Error", "Masukkan angka yang valid.")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = CelengankuGUI()
    window.show()
    sys.exit(app.exec_())
